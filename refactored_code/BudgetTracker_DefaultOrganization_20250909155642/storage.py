import logging
import os
from typing import List
from datetime import date

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.worksheet.worksheet import Worksheet
    from openpyxl.utils import get_column_letter
except ImportError as e:
    raise ImportError("openpyxl is required. Install via: pip install openpyxl") from e

from models import Transaction, BudgetModel

logger = logging.getLogger(__name__)

# Constants
DEFAULT_CATEGORIES = [
    "General",
    "Food",
    "Transport",
    "Housing",
    "Utilities",
    "Entertainment",
    "Healthcare",
    "Income:Salary",
    "Income:Other",
    "Savings",
]
SHEET_TRANSACTIONS = "Transactions"
SHEET_CATEGORIES = "Categories"
SHEET_SETTINGS = "Settings"
SETTINGS_KEY_SAVINGS_GOAL = "savings_goal"
DEFAULT_SAVINGS_GOAL = 0.0
MIN_COLUMN_WIDTH = 10
MAX_COLUMN_WIDTH = 40
COLUMN_WIDTH_PADDING = 2
TRANSACTION_HEADERS = ["ID", "Date", "Type", "Category", "Description", "Amount"]
CATEGORY_HEADERS = ["Name"]
SETTINGS_HEADERS = ["Key", "Value"]


class ExcelStorage:
    """
    Manages Excel-based storage for Budget Tracker using openpyxl.
    
    Handles creation, loading, and saving of workbooks with sheets for
    Transactions, Categories, and Settings.
    """

    def __init__(self, filepath: str):
        """
        Initialize ExcelStorage with the given filepath.
        
        Creates the directory if it doesn't exist and initializes an empty
        file if one doesn't already exist.
        
        Args:
            filepath: Path to the Excel file for storage.
        """
        self.filepath = filepath
        self._ensure_directory_exists()
        if not os.path.exists(self.filepath):
            self._initialize_empty_file()

    def _ensure_directory_exists(self) -> None:
        """Create the directory for the storage file if it doesn't exist."""
        directory = os.path.dirname(os.path.abspath(self.filepath))
        if directory and not os.path.exists(directory):
            os.makedirs(directory, exist_ok=True)
            logger.debug(f"Created directory: {directory}")

    def _initialize_empty_file(self) -> None:
        """
        Create a new Excel file with default sheets and data.
        
        Initializes Transactions, Categories, and Settings sheets with
        appropriate headers and default values.
        """
        wb = Workbook()
        
        # Initialize Transactions sheet
        ws_trans = wb.active
        ws_trans.title = SHEET_TRANSACTIONS
        self._write_sheet_headers(ws_trans, TRANSACTION_HEADERS)

        # Initialize Categories sheet
        ws_cat = wb.create_sheet(SHEET_CATEGORIES)
        self._write_sheet_headers(ws_cat, CATEGORY_HEADERS)
        for category_name in DEFAULT_CATEGORIES:
            ws_cat.append([category_name])

        # Initialize Settings sheet
        ws_settings = wb.create_sheet(SHEET_SETTINGS)
        self._write_sheet_headers(ws_settings, SETTINGS_HEADERS)
        ws_settings.append([SETTINGS_KEY_SAVINGS_GOAL, DEFAULT_SAVINGS_GOAL])

        # Format all sheets
        self._autosize_columns(ws_trans)
        self._autosize_columns(ws_cat)
        self._autosize_columns(ws_settings)
        
        wb.save(self.filepath)
        logger.info(f"Initialized new Excel file: {self.filepath}")

    @staticmethod
    def _write_sheet_headers(ws: Worksheet, headers: List[str]) -> None:
        """
        Write headers to a worksheet.
        
        Args:
            ws: The worksheet to write headers to.
            headers: List of header column names.
        """
        ws.append(headers)

    @staticmethod
    def _autosize_columns(ws: Worksheet) -> None:
        """
        Automatically adjust column widths based on content.
        
        Calculates the maximum content length in each column and sets
        the width accordingly, with bounds between MIN_COLUMN_WIDTH and
        MAX_COLUMN_WIDTH.
        
        Args:
            ws: The worksheet to adjust column widths for.
        """
        # Calculate maximum content length per column
        column_widths = {}
        for row in ws.iter_rows(values_only=True):
            for column_index, cell_value in enumerate(row, start=1):
                cell_text = "" if cell_value is None else str(cell_value)
                current_width = column_widths.get(column_index, 0)
                column_widths[column_index] = max(current_width, len(cell_text))

        # Apply calculated widths with padding and bounds
        for column_index, width in column_widths.items():
            adjusted_width = min(
                max(width + COLUMN_WIDTH_PADDING, MIN_COLUMN_WIDTH),
                MAX_COLUMN_WIDTH
            )
            column_letter = get_column_letter(column_index)
            ws.column_dimensions[column_letter].width = adjusted_width

    def load_model(self) -> BudgetModel:
        """
        Load a BudgetModel from the Excel file.
        
        Reads transactions, categories, and settings from the workbook
        and constructs a BudgetModel instance. Skips malformed rows
        with appropriate logging.
        
        Returns:
            A BudgetModel instance populated with data from the file.
        """
        if not os.path.exists(self.filepath):
            self._initialize_empty_file()

        wb = load_workbook(self.filepath, data_only=True)
        
        transactions = self._load_transactions(wb)
        categories = self._load_categories(wb)
        savings_goal = self._load_savings_goal(wb)

        model = BudgetModel()
        model.load(
            transactions=transactions,
            categories=categories,
            savings_goal=savings_goal
        )
        logger.info(f"Loaded model from {self.filepath}")
        return model

    def _load_transactions(self, wb: Workbook) -> List[Transaction]:
        """
        Load transactions from the Transactions sheet.
        
        Args:
            wb: The loaded workbook.
            
        Returns:
            List of Transaction objects. Malformed rows are skipped.
        """
        transactions = []
        
        if SHEET_TRANSACTIONS not in wb.sheetnames:
            return transactions
        
        trans_sheet = wb[SHEET_TRANSACTIONS]
        is_header_row = True
        
        for row in trans_sheet.iter_rows(values_only=True):
            # Skip header row
            if is_header_row:
                is_header_row = False
                continue
            
            # Skip empty rows
            if not row or row[0] is None:
                continue
            
            try:
                transaction = Transaction.from_row(row)
                transactions.append(transaction)
            except ValueError as e:
                logger.warning(f"Skipped malformed transaction row: {row}. Error: {e}")
            except Exception as e:
                logger.warning(f"Unexpected error loading transaction row {row}: {e}")
        
        return transactions

    def _load_categories(self, wb: Workbook) -> List[str]:
        """
        Load categories from the Categories sheet.
        
        Args:
            wb: The loaded workbook.
            
        Returns:
            List of category names.
        """
        categories = []
        
        if SHEET_CATEGORIES not in wb.sheetnames:
            return categories
        
        cat_sheet = wb[SHEET_CATEGORIES]
        is_header_row = True
        
        for row in cat_sheet.iter_rows(values_only=True):
            # Skip header row
            if is_header_row:
                is_header_row = False
                continue
            
            if row and row[0]:
                categories.append(str(row[0]).strip())
        
        return categories

    def _load_savings_goal(self, wb: Workbook) -> float:
        """
        Load the savings goal from the Settings sheet.
        
        Args:
            wb: The loaded workbook.
            
        Returns:
            The savings goal value, or DEFAULT_SAVINGS_GOAL if not found.
        """
        if SHEET_SETTINGS not in wb.sheetnames:
            return DEFAULT_SAVINGS_GOAL
        
        settings_sheet = wb[SHEET_SETTINGS]
        is_header_row = True
        
        for row in settings_sheet.iter_rows(values_only=True):
            # Skip header row
            if is_header_row:
                is_header_row = False
                continue
            
            if row and row[0] == SETTINGS_KEY_SAVINGS_GOAL:
                try:
                    return float(row[1])
                except (ValueError, TypeError) as e:
                    logger.warning(f"Invalid savings_goal value: {row[1]}. Using default.")
                    return DEFAULT_SAVINGS_GOAL
        
        return DEFAULT_SAVINGS_GOAL

    def _build_workbook(self, model: BudgetModel) -> Workbook:
        """
        Build an Excel workbook from the provided model without writing to disk.
        
        This method encapsulates the workbook creation logic to avoid duplication
        between save_model and export_copy.
        
        Args:
            model: The BudgetModel to convert to a workbook.
            
        Returns:
            A Workbook instance with all model data populated.
        """
        wb = Workbook()

        # Build Transactions sheet
        ws_trans = wb.active
        ws_trans.title = SHEET_TRANSACTIONS
        self._write_sheet_headers(ws_trans, TRANSACTION_HEADERS)
        for transaction in model.transactions:
            row = list(transaction.to_row())
            # Store amount as positive; sign is inferred from Type field
            row[5] = float(abs(row[5]))
            ws_trans.append(row)

        # Build Categories sheet
        ws_cat = wb.create_sheet(SHEET_CATEGORIES)
        self._write_sheet_headers(ws_cat, CATEGORY_HEADERS)
        for category_name in model.categories:
            ws_cat.append([category_name])

        # Build Settings sheet
        ws_settings = wb.create_sheet(SHEET_SETTINGS)
        self._write_sheet_headers(ws_settings, SETTINGS_HEADERS)
        ws_settings.append([SETTINGS_KEY_SAVINGS_GOAL, float(model.savings_goal)])

        # Format all sheets
        self._autosize_columns(ws_trans)
        self._autosize_columns(ws_cat)
        self._autosize_columns(ws_settings)

        return wb

    def save_model(self, model: BudgetModel) -> None:
        """
        Save the BudgetModel to the Excel file.
        
        Creates a fresh workbook and writes all model content to the
        configured storage filepath.
        
        Args:
            model: The BudgetModel to save.
        """
        wb = self._build_workbook(model)
        wb.save(self.filepath)
        logger.info(f"Saved model to {self.filepath}")

    def export_copy(self, model: BudgetModel, dest_path: str) -> None:
        """
        Export a copy of the model to an arbitrary destination.
        
        Creates a new workbook without modifying the active storage filepath.
        
        Args:
            model: The BudgetModel to export.
            dest_path: The destination filepath for the export.
        """
        wb = self._build_workbook(model)
        wb.save(dest_path)
        logger.info(f"Exported model copy to {dest_path}")